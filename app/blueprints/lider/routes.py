from flask import render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from app.blueprints.lider import lider_bp
from app.blueprints.auth.routes import role_required
from app.models.solicitud import Solicitud
from app.models.servicio import Servicio
from app.models.ubicacion import Ubicacion
from app.models.usuario import Usuario
from app.models.localidad import Localidad
from app.models.vehiculo import Vehiculo
from app import db
from datetime import datetime, timedelta
from sqlalchemy import func, and_, or_, text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import pandas as pd

@lider_bp.route('/dashboard')
@login_required
@role_required('lider')
def dashboard():
    """Dashboard principal del líder"""
    # Métricas generales
    hoy = datetime.utcnow().date()
    inicio_mes = datetime(hoy.year, hoy.month, 1)
    
    # Solicitudes del día
    # Usar comparación de rango de fechas para SQLite
    inicio_dia = datetime.combine(hoy, datetime.min.time())
    fin_dia = datetime.combine(hoy, datetime.max.time())
    solicitudes_hoy = Solicitud.query.filter(
        Solicitud.created_at.between(inicio_dia, fin_dia)
    ).count()
    
    # Servicios activos
    servicios_activos = Servicio.query.filter(
        Servicio.estado_servicio.in_(['aceptado', 'en_ruta', 'en_sitio'])
    ).count()
    
    # Móviles activas (con actividad reciente)
    tiempo_limite = datetime.utcnow() - timedelta(minutes=30)
    moviles_activas = Usuario.query.filter(
        Usuario.rol == 'movil',
        Usuario.activo == True,
        Usuario.last_activity >= tiempo_limite
    ).count()
    
    # Solicitudes pendientes
    solicitudes_pendientes = Solicitud.query.filter_by(estado='pendiente').count()
    
    # Métricas del mes
    solicitudes_mes = Solicitud.query.filter(
        Solicitud.created_at >= inicio_mes
    ).count()
    
    servicios_completados_mes = Servicio.query.filter(
        Servicio.estado_servicio == 'completado',
        Servicio.finalizado_at >= inicio_mes
    ).count()
    
    # Tasa de aceptación del mes
    solicitudes_aceptadas_mes = Solicitud.query.filter(
        Solicitud.estado.in_(['aceptada', 'completada']),
        Solicitud.created_at >= inicio_mes
    ).count()
    
    tasa_aceptacion = 0
    if solicitudes_mes > 0:
        tasa_aceptacion = round((solicitudes_aceptadas_mes / solicitudes_mes) * 100, 1)
    
    # Tiempo promedio de respuesta (en minutos)
    # Usar funciones de MySQL para calcular diferencia en minutos
    servicios_con_tiempo = db.session.query(
        func.avg(
            func.timestampdiff(text('MINUTE'), Solicitud.created_at, Servicio.aceptado_at)
        )
    ).select_from(Servicio).join(
        Solicitud, Servicio.solicitud_id == Solicitud.id
    ).filter(
        Servicio.aceptado_at >= inicio_mes
    ).scalar()
    
    tiempo_promedio_respuesta = round(servicios_con_tiempo or 0, 1)
    
    # Solicitudes recientes
    solicitudes_recientes = Solicitud.query.order_by(
        Solicitud.created_at.desc()
    ).limit(10).all()
    
    # Servicios activos detallados
    servicios_activos_detalle = Servicio.query.filter(
        Servicio.estado_servicio.in_(['aceptado', 'en_ruta', 'en_sitio'])
    ).order_by(Servicio.aceptado_at.desc()).all()
    
    return render_template('lider/dashboard.html',
                         solicitudes_hoy=solicitudes_hoy,
                         servicios_activos=servicios_activos,
                         moviles_activas=moviles_activas,
                         solicitudes_pendientes=solicitudes_pendientes,
                         solicitudes_mes=solicitudes_mes,
                         servicios_completados_mes=servicios_completados_mes,
                         tasa_aceptacion=tasa_aceptacion,
                         tiempo_promedio_respuesta=tiempo_promedio_respuesta,
                         solicitudes_recientes=solicitudes_recientes,
                         servicios_activos_detalle=servicios_activos_detalle)

@lider_bp.route('/mapa')
@login_required
@role_required('lider')
def mapa():
    """Mapa general con todas las operaciones"""
    # Obtener todas las ubicaciones activas de móviles
    tiempo_limite = datetime.utcnow() - timedelta(minutes=30)
    
    moviles_ubicaciones = db.session.query(Usuario, Ubicacion).join(
        Ubicacion, Usuario.id == Ubicacion.usuario_id
    ).filter(
        Usuario.rol == 'movil',
        Usuario.activo == True,
        Ubicacion.activa == True,
        Usuario.last_activity >= tiempo_limite
    ).all()
    
    # Solicitudes pendientes y activas
    solicitudes_activas = Solicitud.query.filter(
        Solicitud.estado.in_(['pendiente', 'aceptada'])
    ).all()
    
    # Servicios en curso
    servicios_activos = Servicio.query.filter(
        Servicio.estado_servicio.in_(['aceptado', 'en_ruta', 'en_sitio'])
    ).all()
    
    return render_template('lider/mapa.html',
                         moviles_ubicaciones=moviles_ubicaciones,
                         solicitudes_activas=solicitudes_activas,
                         servicios_activos=servicios_activos)

@lider_bp.route('/reportes')
@login_required
@role_required('lider')
def reportes():
    """Página de reportes y métricas"""
    # Parámetros de filtro
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    
    # Fechas por defecto (último mes)
    if not fecha_inicio or not fecha_fin:
        fecha_fin = datetime.utcnow().date()
        fecha_inicio = fecha_fin - timedelta(days=30)
    else:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    
    # Convertir a datetime para consultas
    inicio_dt = datetime.combine(fecha_inicio, datetime.min.time())
    fin_dt = datetime.combine(fecha_fin, datetime.max.time())
    
    # Métricas del período
    total_solicitudes = Solicitud.query.filter(
        Solicitud.created_at.between(inicio_dt, fin_dt)
    ).count()
    
    solicitudes_aceptadas = Solicitud.query.filter(
        Solicitud.created_at.between(inicio_dt, fin_dt),
        Solicitud.estado.in_(['aceptada', 'completada'])
    ).count()
    
    solicitudes_rechazadas = Solicitud.query.filter(
        Solicitud.created_at.between(inicio_dt, fin_dt),
        Solicitud.estado == 'rechazada'
    ).count()
    
    solicitudes_expiradas = Solicitud.query.filter(
        Solicitud.created_at.between(inicio_dt, fin_dt),
        Solicitud.estado == 'expirada'
    ).count()
    
    servicios_completados = db.session.query(Servicio).select_from(Servicio).join(
        Solicitud, Servicio.solicitud_id == Solicitud.id
    ).filter(
        Solicitud.created_at.between(inicio_dt, fin_dt),
        Servicio.estado_servicio == 'completado'
    ).count()
    
    # Rendimiento por móvil
    # Obtener datos básicos de móviles
    rendimiento_moviles_raw = db.session.query(
        Usuario.id,
        Usuario.nombre,
        Usuario.apellido,
        func.count(Servicio.id).label('total_servicios'),
        func.avg(Servicio.duracion_minutos).label('duracion_promedio')
    ).select_from(Usuario).join(
        Servicio, Usuario.id == Servicio.movil_id
    ).join(
        Solicitud, Servicio.solicitud_id == Solicitud.id
    ).filter(
        Usuario.rol == 'movil',
        Solicitud.created_at.between(inicio_dt, fin_dt)
    ).group_by(Usuario.id).all()
    
    # Calcular servicios completados por separado
    rendimiento_moviles = []
    for movil in rendimiento_moviles_raw:
        completados = db.session.query(func.count(Servicio.id)).select_from(Servicio).join(
            Solicitud, Servicio.solicitud_id == Solicitud.id
        ).filter(
            Servicio.movil_id == movil.id,
            Servicio.estado_servicio == 'completado',
            Solicitud.created_at.between(inicio_dt, fin_dt)
        ).scalar() or 0
        
        rendimiento_moviles.append({
            'nombre': movil.nombre,
            'apellido': movil.apellido,
            'total_servicios': movil.total_servicios,
            'completados': completados,
            'duracion_promedio': movil.duracion_promedio
        })
    
    # Tipos de apoyo más solicitados
    tipos_apoyo_raw = db.session.query(
        Solicitud.tipo_apoyo,
        func.count(Solicitud.id).label('cantidad')
    ).filter(
        Solicitud.created_at.between(inicio_dt, fin_dt)
    ).group_by(Solicitud.tipo_apoyo).all()
    
    tipos_apoyo = [{
        'tipo_apoyo': row.tipo_apoyo,
        'cantidad': row.cantidad
    } for row in tipos_apoyo_raw]
    
    # Solicitudes por día
    # Usar DATE_FORMAT para MySQL
    solicitudes_por_dia_raw = db.session.query(
        func.date_format(Solicitud.created_at, '%Y-%m-%d').label('fecha'),
        func.count(Solicitud.id).label('cantidad')
    ).filter(
        Solicitud.created_at.between(inicio_dt, fin_dt)
    ).group_by(func.date_format(Solicitud.created_at, '%Y-%m-%d')).order_by('fecha').all()
    
    solicitudes_por_dia = [{
        'fecha': row.fecha,
        'cantidad': row.cantidad
    } for row in solicitudes_por_dia_raw]
    
    return render_template('lider/reportes.html',
                         fecha_inicio=fecha_inicio,
                         fecha_fin=fecha_fin,
                         total_solicitudes=total_solicitudes,
                         solicitudes_aceptadas=solicitudes_aceptadas,
                         solicitudes_rechazadas=solicitudes_rechazadas,
                         solicitudes_expiradas=solicitudes_expiradas,
                         servicios_completados=servicios_completados,
                         rendimiento_moviles=rendimiento_moviles,
                         tipos_apoyo=tipos_apoyo,
                         solicitudes_por_dia=solicitudes_por_dia)

@lider_bp.route('/reportes/exportar-excel')
@login_required
@role_required('lider')
def exportar_reportes_excel():
    """Exportar reportes de líder a Excel"""
    try:
        # Parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        # Fechas por defecto (último mes)
        if not fecha_inicio or not fecha_fin:
            fecha_fin = datetime.utcnow().date()
            fecha_inicio = fecha_fin - timedelta(days=30)
        else:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        
        # Convertir a datetime para consultas
        inicio_dt = datetime.combine(fecha_inicio, datetime.min.time())
        fin_dt = datetime.combine(fecha_fin, datetime.max.time())
        
        # Crear workbook
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === HOJA 1: RESUMEN GENERAL ===
        ws_resumen = wb.active
        ws_resumen.title = "Resumen General"
        
        # Título principal
        ws_resumen['A1'] = f"REPORTE DE LÍDER - SYNAPSIS APOYOS"
        ws_resumen['A1'].font = Font(bold=True, size=16)
        ws_resumen['A2'] = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        ws_resumen['A2'].font = Font(bold=True, size=12)
        ws_resumen['A3'] = f"Generado: {datetime.utcnow().strftime('%d/%m/%Y %H:%M')}"
        
        # Métricas generales
        total_solicitudes = Solicitud.query.filter(
            Solicitud.created_at.between(inicio_dt, fin_dt)
        ).count()
        
        solicitudes_aceptadas = Solicitud.query.filter(
            Solicitud.created_at.between(inicio_dt, fin_dt),
            Solicitud.estado.in_(['aceptada', 'completada'])
        ).count()
        
        solicitudes_rechazadas = Solicitud.query.filter(
            Solicitud.created_at.between(inicio_dt, fin_dt),
            Solicitud.estado == 'rechazada'
        ).count()
        
        servicios_completados = db.session.query(Servicio).select_from(Servicio).join(
            Solicitud, Servicio.solicitud_id == Solicitud.id
        ).filter(
            Solicitud.created_at.between(inicio_dt, fin_dt),
            Servicio.estado_servicio == 'completado'
        ).count()
        
        # Escribir métricas
        row = 5
        metricas = [
            ('MÉTRICA', 'VALOR'),
            ('Total Solicitudes', total_solicitudes),
            ('Solicitudes Aceptadas', solicitudes_aceptadas),
            ('Solicitudes Rechazadas', solicitudes_rechazadas),
            ('Servicios Completados', servicios_completados),
            ('Tasa de Aceptación (%)', round((solicitudes_aceptadas / total_solicitudes * 100) if total_solicitudes > 0 else 0, 2))
        ]
        
        for i, (metrica, valor) in enumerate(metricas):
            ws_resumen[f'A{row + i}'] = metrica
            ws_resumen[f'B{row + i}'] = valor
            if i == 0:  # Header
                ws_resumen[f'A{row + i}'].font = header_font
                ws_resumen[f'A{row + i}'].fill = header_fill
                ws_resumen[f'B{row + i}'].font = header_font
                ws_resumen[f'B{row + i}'].fill = header_fill
            ws_resumen[f'A{row + i}'].border = border
            ws_resumen[f'B{row + i}'].border = border
        
        # === HOJA 2: SERVICIOS DETALLADOS ===
        ws_servicios = wb.create_sheet("Servicios Detallados")
        
        # Obtener servicios del período
        servicios = db.session.query(Servicio, Solicitud, Usuario).select_from(Servicio).join(
            Solicitud, Servicio.solicitud_id == Solicitud.id
        ).join(
            Usuario, Servicio.movil_id == Usuario.id, isouter=True
        ).filter(
            Solicitud.created_at.between(inicio_dt, fin_dt)
        ).all()
        
        # Headers para servicios
        headers_servicios = [
            'ID Servicio', 'Estado', 'Tipo Apoyo', 'Dirección', 'Coordenadas',
            'Móvil Asignado', 'Fecha Creación', 'Fecha Aceptación', 'Fecha Finalización',
            'Duración (min)', 'Observaciones'
        ]
        
        for col, header in enumerate(headers_servicios, 1):
            cell = ws_servicios.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos de servicios
        for row_idx, (servicio, solicitud, movil) in enumerate(servicios, 2):
            data_row = [
                servicio.id,
                servicio.estado_servicio,
                solicitud.tipo_apoyo,
                solicitud.direccion,
                f"{solicitud.latitud}, {solicitud.longitud}",
                f"{movil.nombre} {movil.apellido}" if movil else "No asignado",
                solicitud.created_at.strftime('%d/%m/%Y %H:%M'),
                servicio.aceptado_at.strftime('%d/%m/%Y %H:%M') if servicio.aceptado_at else '',
                servicio.finalizado_at.strftime('%d/%m/%Y %H:%M') if servicio.finalizado_at else '',
                servicio.duracion_minutos or '',
                servicio.observaciones or ''
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws_servicios.cell(row=row_idx, column=col, value=value)
                cell.border = border
        
        # === HOJA 3: RENDIMIENTO POR MÓVIL ===
        ws_moviles = wb.create_sheet("Rendimiento Móviles")
        
        # Obtener rendimiento por móvil
        rendimiento_moviles_raw = db.session.query(
            Usuario.id,
            Usuario.nombre,
            Usuario.apellido,
            func.count(Servicio.id).label('total_servicios'),
            func.avg(Servicio.duracion_minutos).label('duracion_promedio')
        ).select_from(Usuario).join(
            Servicio, Usuario.id == Servicio.movil_id
        ).join(
            Solicitud, Servicio.solicitud_id == Solicitud.id
        ).filter(
            Usuario.rol == 'movil',
            Solicitud.created_at.between(inicio_dt, fin_dt)
        ).group_by(Usuario.id).all()
        
        # Headers para móviles
        headers_moviles = [
            'Móvil', 'Total Servicios', 'Servicios Completados', 'Tasa Éxito (%)', 'Tiempo Promedio (min)'
        ]
        
        for col, header in enumerate(headers_moviles, 1):
            cell = ws_moviles.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Datos de móviles
        for row_idx, movil in enumerate(rendimiento_moviles_raw, 2):
            completados = db.session.query(func.count(Servicio.id)).select_from(Servicio).join(
                Solicitud, Servicio.solicitud_id == Solicitud.id
            ).filter(
                Servicio.movil_id == movil.id,
                Servicio.estado_servicio == 'completado',
                Solicitud.created_at.between(inicio_dt, fin_dt)
            ).scalar() or 0
            
            tasa_exito = round((completados / movil.total_servicios * 100) if movil.total_servicios > 0 else 0, 2)
            
            data_row = [
                f"{movil.nombre} {movil.apellido}",
                movil.total_servicios,
                completados,
                tasa_exito,
                round(movil.duracion_promedio, 2) if movil.duracion_promedio else 0
            ]
            
            for col, value in enumerate(data_row, 1):
                cell = ws_moviles.cell(row=row_idx, column=col, value=value)
                cell.border = border
        
        # Ajustar ancho de columnas
        for ws in [ws_resumen, ws_servicios, ws_moviles]:
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Nombre del archivo con timestamp
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        filename = f'reporte_lider_{timestamp}.xlsx'
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al generar reporte Excel: {str(e)}'
        }), 500

@lider_bp.route('/api/localidades')
@login_required
@role_required('lider')
def obtener_localidades():
    """Obtener lista de localidades activas"""
    try:
        localidades = Localidad.get_all_active()
        localidades_data = []
        
        for localidad in localidades:
            localidades_data.append({
                'codigo': localidad.codigo,
                'nombre': localidad.nombre
            })
        
        # Ordenar por nombre
        localidades_data.sort(key=lambda x: x['nombre'])
        
        return jsonify({
            'success': True,
            'localidades': localidades_data
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al obtener localidades: {str(e)}'
        }), 500

@lider_bp.route('/solicitudes')
@login_required
@role_required('lider')
def solicitudes():
    """Lista de todas las solicitudes"""
    page = request.args.get('page', 1, type=int)
    estado = request.args.get('estado', 'todas')
    tecnico_id = request.args.get('tecnico_id', type=int)
    localidad_codigo = request.args.get('localidad', 'todas')
    tipo_apoyo = request.args.get('tipo_apoyo', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    
    query = Solicitud.query
    
    if estado != 'todas':
        query = query.filter_by(estado=estado)
    
    if tecnico_id:
        query = query.filter_by(tecnico_id=tecnico_id)
    
    # Filtrar por localidad usando el campo localidad_id
    if localidad_codigo != 'todas':
        localidad = Localidad.get_by_codigo(localidad_codigo)
        if localidad:
            query = query.filter_by(localidad_id=localidad.id)
        else:
            # Si no se encuentra la localidad, mostrar lista vacía
            query = query.filter(Solicitud.id == -1)
    
    # Filtrar por tipo de apoyo
    if tipo_apoyo:
        query = query.filter_by(tipo_apoyo=tipo_apoyo)
    
    # Filtrar por fechas
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            query = query.filter(Solicitud.created_at >= fecha_desde_dt)
        except ValueError:
            pass  # Ignorar fecha inválida
    
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            # Incluir todo el día hasta las 23:59:59
            fecha_hasta_dt = fecha_hasta_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Solicitud.created_at <= fecha_hasta_dt)
        except ValueError:
            pass  # Ignorar fecha inválida
    
    # Aplicar paginación a la query final
    solicitudes = query.order_by(Solicitud.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Lista de técnicos para filtro
    tecnicos = Usuario.query.filter_by(rol='tecnico', activo=True).all()
    
    return render_template('lider/solicitudes.html',
                         solicitudes=solicitudes,
                         estado_filtro=estado,
                         tecnico_filtro=tecnico_id,
                         localidad_filtro=localidad_codigo,
                         tipo_apoyo_filtro=tipo_apoyo,
                         fecha_desde_filtro=fecha_desde,
                         fecha_hasta_filtro=fecha_hasta,
                         tecnicos=tecnicos)

@lider_bp.route('/api/solicitud/<int:solicitud_id>/detalles')
@login_required
@role_required('lider')
def obtener_detalles_solicitud(solicitud_id):
    """Obtiene los detalles completos de una solicitud"""
    solicitud = Solicitud.query.get_or_404(solicitud_id)
    
    # Obtener información del técnico
    tecnico = Usuario.query.get(solicitud.tecnico_id)
    
    # Obtener información del servicio si existe
    servicio = None
    movil = None
    if solicitud.servicio:
        servicio = solicitud.servicio
        movil = Usuario.query.get(servicio.movil_id) if servicio.movil_id else None
    
    # Obtener observaciones (es un campo de texto simple)
    observaciones = []
    if solicitud.observaciones:
        observaciones.append({
            'id': 1,
            'contenido': solicitud.observaciones,
            'created_at': solicitud.created_at.strftime('%d/%m/%Y %H:%M'),
            'usuario': tecnico.get_nombre_completo() if tecnico else 'Sistema'
        })
    
    # Preparar respuesta
    detalles = {
        'id': solicitud.id,
        'estado': solicitud.estado,
        'tipo_apoyo': solicitud.tipo_apoyo,
        'descripcion': solicitud.observaciones,
        'direccion': solicitud.direccion,
        'latitud': float(solicitud.latitud),
        'longitud': float(solicitud.longitud),
        'created_at': solicitud.created_at.strftime('%d/%m/%Y %H:%M'),
        'updated_at': solicitud.updated_at.strftime('%d/%m/%Y %H:%M') if solicitud.updated_at else None,
        'tecnico': {
            'id': tecnico.id,
            'nombre': tecnico.get_nombre_completo(),
            'telefono': tecnico.telefono,
            'email': tecnico.email
        } if tecnico else None,
        'servicio': {
            'id': servicio.id,
            'estado': servicio.estado_servicio,
            'aceptado_at': servicio.aceptado_at.strftime('%d/%m/%Y %H:%M') if servicio.aceptado_at else None,
            'finalizado_at': servicio.finalizado_at.strftime('%d/%m/%Y %H:%M') if servicio.finalizado_at else None,
            'duracion_minutos': servicio.duracion_minutos,
            'observaciones_finales': servicio.observaciones_finales,
            'movil': {
                'id': movil.id,
                'nombre': movil.get_nombre_completo(),
                'telefono': movil.telefono
            } if movil else None
        } if servicio else None,
        'observaciones': observaciones
    }
    
    return jsonify(detalles)

@lider_bp.route('/servicios')
@login_required
@role_required('lider')
def servicios():
    """Lista de todos los servicios"""
    page = request.args.get('page', 1, type=int)
    estado = request.args.get('estado', 'todos')
    movil_id = request.args.get('movil_id', type=int)
    
    query = Servicio.query
    
    if estado != 'todos':
        query = query.filter_by(estado_servicio=estado)
    
    if movil_id:
        query = query.filter_by(movil_id=movil_id)
    
    servicios = query.order_by(Servicio.aceptado_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # Lista de móviles para filtro
    moviles = Usuario.query.filter_by(rol='movil', activo=True).all()
    
    return render_template('lider/servicios.html',
                         servicios=servicios,
                         estado_filtro=estado,
                         movil_filtro=movil_id,
                         moviles=moviles)

@lider_bp.route('/usuarios')
@login_required
@role_required('lider')
def usuarios():
    """Gestión de usuarios"""
    # Obtener parámetros de filtro
    rol_filtro = request.args.get('rol', 'todos')
    estado_filtro = request.args.get('estado', 'todos')
    buscar_filtro = request.args.get('buscar', '')
    page = request.args.get('page', 1, type=int)
    
    # Calcular estadísticas para las tarjetas
    total_usuarios = Usuario.query.count()
    usuarios_activos = Usuario.query.filter_by(activo=True).count()
    tecnicos_count = Usuario.query.filter_by(rol='tecnico').count()
    moviles_count = Usuario.query.filter_by(rol='movil').count()
    
    # Construir consulta base
    query = Usuario.query
    
    # Aplicar filtro de rol
    if rol_filtro != 'todos':
        query = query.filter_by(rol=rol_filtro)
    
    # Aplicar filtro de estado
    if estado_filtro == 'activo':
        query = query.filter_by(activo=True)
    elif estado_filtro == 'inactivo':
        query = query.filter_by(activo=False)
    
    # Aplicar filtro de búsqueda
    if buscar_filtro:
        search_term = f"%{buscar_filtro}%"
        query = query.filter(
            or_(
                Usuario.nombre.ilike(search_term),
                Usuario.apellido.ilike(search_term),
                Usuario.email.ilike(search_term),
                Usuario.telefono.ilike(search_term)
            )
        )
    
    # Aplicar paginación
    usuarios = query.order_by(Usuario.created_at.desc()).paginate(
        page=page, per_page=10, error_out=False
    )
    
    return render_template('lider/usuarios.html',
                         usuarios=usuarios,
                         rol_filtro=rol_filtro,
                         estado_filtro=estado_filtro,
                         buscar_filtro=buscar_filtro,
                         total_usuarios=total_usuarios,
                         usuarios_activos=usuarios_activos,
                         tecnicos_count=tecnicos_count,
                         moviles_count=moviles_count)

@lider_bp.route('/usuario/crear', methods=['POST'])
@login_required
@role_required('lider')
def crear_usuario():
    """Crear nuevo usuario"""
    try:
        data = request.get_json() if request.is_json else request.form
        
        # Validar datos requeridos
        required_fields = ['nombre', 'apellido', 'email', 'telefono', 'rol', 'password']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'message': f'El campo {field} es requerido'
                }), 400
        
        # Verificar si el email ya existe
        if Usuario.query.filter_by(email=data['email']).first():
            return jsonify({
                'success': False,
                'message': 'El email ya está registrado'
            }), 400
        
        # Crear nuevo usuario
        nuevo_usuario = Usuario(
            email=data['email'],
            password=data['password'],
            nombre=data['nombre'],
            apellido=data['apellido'],
            rol=data['rol'],
            telefono=data['telefono']
        )
        
        db.session.add(nuevo_usuario)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Usuario {nuevo_usuario.get_nombre_completo()} creado exitosamente',
            'usuario': nuevo_usuario.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error creando usuario: {str(e)}'
        }), 500

@lider_bp.route('/usuario/<int:usuario_id>/editar', methods=['POST'])
@login_required
@role_required('lider')
def editar_usuario(usuario_id):
    """Editar usuario existente"""
    try:
        usuario = Usuario.query.get_or_404(usuario_id)
        data = request.get_json() if request.is_json else request.form
        
        # Actualizar campos si se proporcionan
        if data.get('nombre'):
            usuario.nombre = data['nombre']
        if data.get('apellido'):
            usuario.apellido = data['apellido']
        if data.get('email'):
            # Verificar que el email no esté en uso por otro usuario
            email_existente = Usuario.query.filter(
                Usuario.email == data['email'],
                Usuario.id != usuario_id
            ).first()
            if email_existente:
                return jsonify({
                    'success': False,
                    'message': 'El email ya está en uso por otro usuario'
                }), 400
            usuario.email = data['email']
        if data.get('telefono'):
            usuario.telefono = data['telefono']
        if data.get('rol'):
            usuario.rol = data['rol']
        if data.get('password'):
            usuario.set_password(data['password'])
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Usuario {usuario.get_nombre_completo()} actualizado exitosamente',
            'usuario': usuario.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error actualizando usuario: {str(e)}'
        }), 500

@lider_bp.route('/usuario/<int:usuario_id>/perfil', methods=['GET'])
@login_required
@role_required('lider')
def ver_perfil_usuario(usuario_id):
    """Ver perfil detallado de usuario"""
    try:
        usuario = Usuario.query.get_or_404(usuario_id)
        
        # Obtener estadísticas del usuario según su rol
        estadisticas = {}
        
        if usuario.rol == 'tecnico':
            # Estadísticas para técnicos
            total_solicitudes = Solicitud.query.filter_by(tecnico_id=usuario.id).count()
            solicitudes_aceptadas = Solicitud.query.filter_by(
                tecnico_id=usuario.id, estado='aceptada'
            ).count()
            
            estadisticas = {
                'total_solicitudes': total_solicitudes,
                'solicitudes_aceptadas': solicitudes_aceptadas,
                'tasa_aceptacion': round((solicitudes_aceptadas / total_solicitudes * 100) if total_solicitudes > 0 else 0, 1)
            }
            
        elif usuario.rol == 'movil':
            # Estadísticas para móviles
            total_servicios = Servicio.query.filter_by(movil_id=usuario.id).count()
            servicios_completados = Servicio.query.filter_by(
                movil_id=usuario.id, estado_servicio='completado'
            ).count()
            
            estadisticas = {
                'total_servicios': total_servicios,
                'servicios_completados': servicios_completados,
                'tasa_completacion': round((servicios_completados / total_servicios * 100) if total_servicios > 0 else 0, 1)
            }
        
        return jsonify({
            'success': True,
            'usuario': usuario.to_dict(),
            'estadisticas': estadisticas
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error obteniendo perfil: {str(e)}'
        }), 500

@lider_bp.route('/usuario/<int:usuario_id>/resetear-password', methods=['POST'])
@login_required
@role_required('lider')
def resetear_password(usuario_id):
    """Resetear contraseña de usuario y enviar por correo"""
    try:
        usuario = Usuario.query.get_or_404(usuario_id)
        
        # Verificar que el usuario tenga email
        if not usuario.email:
            return jsonify({
                'success': False,
                'message': f'El usuario {usuario.get_nombre_completo()} no tiene un email registrado'
            }), 400
        
        # Generar contraseña temporal segura
        from app.utils.password_utils import generate_temporary_password
        nueva_password = generate_temporary_password()
        
        # Actualizar la contraseña en la base de datos
        usuario.set_password(nueva_password)
        db.session.commit()
        
        # Enviar correo con la nueva contraseña
        from app.utils.email_utils import send_password_reset_email
        email_enviado, mensaje_email = send_password_reset_email(usuario, nueva_password)
        
        if email_enviado:
            return jsonify({
                'success': True,
                'message': f'Contraseña reseteada para {usuario.get_nombre_completo()}. Se ha enviado un correo a {usuario.email} con los detalles.',
                'email_enviado': True
            })
        else:
            # Si el correo falló, aún así la contraseña fue cambiada
            return jsonify({
                'success': True,
                'message': f'Contraseña reseteada para {usuario.get_nombre_completo()}, pero hubo un error al enviar el correo: {mensaje_email}',
                'email_enviado': False,
                'nueva_password': nueva_password  # Solo mostrar si el correo falló
            })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error reseteando contraseña: {str(e)}'
        }), 500

@lider_bp.route('/usuario/<int:usuario_id>/toggle-activo', methods=['POST'])
@login_required
@role_required('lider')
def toggle_usuario_activo(usuario_id):
    """Activar/desactivar usuario"""
    try:
        usuario = Usuario.query.get_or_404(usuario_id)
        
        if usuario.id == current_user.id:
            flash('No puedes desactivar tu propia cuenta', 'error')
            return redirect(url_for('lider.usuarios'))
        
        usuario.activo = not usuario.activo
        db.session.commit()
        
        estado = 'activado' if usuario.activo else 'desactivado'
        flash(f'Usuario {usuario.get_nombre_completo()} {estado}', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error actualizando usuario: {str(e)}', 'error')
    
    return redirect(url_for('lider.usuarios'))

@lider_bp.route('/perfil')
@login_required
@role_required('lider')
def perfil():
    """Perfil del líder"""
    return render_template('lider/perfil.html')

@lider_bp.route('/perfil/actualizar-informacion', methods=['POST'])
@login_required
@role_required('lider')
def actualizar_informacion():
    """Actualizar información personal del líder"""
    try:
        data = request.get_json()
        
        # Actualizar campos si se proporcionan
        if data.get('nombre'):
            current_user.nombre = data['nombre']
        if data.get('apellido'):
            current_user.apellido = data['apellido']
        if data.get('email'):
            # Verificar que el email no esté en uso por otro usuario
            email_existente = Usuario.query.filter(
                Usuario.email == data['email'],
                Usuario.id != current_user.id
            ).first()
            if email_existente:
                return jsonify({
                    'success': False,
                    'message': 'El email ya está en uso por otro usuario'
                }), 400
            current_user.email = data['email']
        if data.get('telefono'):
            current_user.telefono = data['telefono']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Información actualizada correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error actualizando información: {str(e)}'
        }), 500

@lider_bp.route('/perfil/cambiar-password', methods=['POST'])
@login_required
@role_required('lider')
def cambiar_password():
    """Cambiar contraseña del líder"""
    try:
        data = request.get_json()
        
        password_actual = data.get('password_actual')
        password_nueva = data.get('password_nueva')
        password_confirmar = data.get('password_confirmar')
        
        # Validaciones
        if not password_actual or not password_nueva or not password_confirmar:
            return jsonify({
                'success': False,
                'message': 'Todos los campos de contraseña son requeridos'
            }), 400
        
        if not current_user.check_password(password_actual):
            return jsonify({
                'success': False,
                'message': 'La contraseña actual es incorrecta'
            }), 400
        
        if password_nueva != password_confirmar:
            return jsonify({
                'success': False,
                'message': 'Las contraseñas no coinciden'
            }), 400
        
        if len(password_nueva) < 6:
            return jsonify({
                'success': False,
                'message': 'La nueva contraseña debe tener al menos 6 caracteres'
            }), 400
        
        # Actualizar contraseña
        current_user.set_password(password_nueva)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Contraseña cambiada correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error cambiando contraseña: {str(e)}'
        }), 500

# ============================================================================
# APIs para el mapa en tiempo real del líder
# ============================================================================

@lider_bp.route('/api/moviles-tiempo-real')
@login_required
@role_required('lider')
def api_moviles_tiempo_real():
    """API para obtener móviles en tiempo real para el mapa del líder"""
    try:
        moviles = []
        
        # Consultar móviles con sesiones activas (últimos 30 minutos)
        tiempo_limite = datetime.utcnow() - timedelta(minutes=30)
        
        # Obtener usuarios móviles con actividad reciente
        moviles_query = db.session.query(Usuario, Ubicacion).join(
            Ubicacion, Usuario.id == Ubicacion.usuario_id
        ).filter(
            Usuario.rol == 'movil',
            Usuario.activo == True,
            Usuario.last_activity >= tiempo_limite,
            Ubicacion.activa == True
        ).all()
        
        for usuario, ubicacion in moviles_query:
            # Determinar estado de la móvil
            servicio_activo = Servicio.query.filter(
                Servicio.movil_id == usuario.id,
                Servicio.estado_servicio.in_(['aceptado', 'en_ruta', 'en_sitio'])
            ).first()
            
            if servicio_activo:
                estado = 'ocupada'
                estado_color = '#dc3545'  # rojo
                estado_texto = f'En servicio ({servicio_activo.estado_servicio})'
                servicio_actual = f'Servicio #{servicio_activo.id}'
            else:
                estado = 'disponible'
                estado_color = '#28a745'  # verde
                estado_texto = 'Disponible'
                servicio_actual = None
            
            moviles.append({
                'id': usuario.id,
                'nombre': usuario.get_nombre_completo(),
                'coordenadas': {
                    'lat': float(ubicacion.latitud),
                    'lng': float(ubicacion.longitud)
                },
                'estado': estado,
                'estado_color': estado_color,
                'estado_texto': estado_texto,
                'telefono': usuario.telefono or 'No disponible',
                'velocidad_kmh': 0,  # Campo no disponible en el modelo
                'direccion_movimiento': 0,  # Campo no disponible en el modelo
                'servicio_actual': servicio_actual,
                'ultima_actualizacion': ubicacion.timestamp.strftime('%H:%M:%S')
            })
        
        return jsonify({
            'success': True,
            'moviles': moviles
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@lider_bp.route('/api/solicitudes-tiempo-real')
@login_required
@role_required('lider')
def api_solicitudes_tiempo_real():
    """API para obtener solicitudes en tiempo real para el mapa del líder"""
    try:
        solicitudes = []
        
        # Consultar solicitudes pendientes y aceptadas
        solicitudes_query = Solicitud.query.filter(
            Solicitud.estado.in_(['pendiente', 'aceptada'])
        ).all()
        
        for solicitud in solicitudes_query:
            # Obtener información del técnico
            tecnico = Usuario.query.get(solicitud.tecnico_id)
            
            # Determinar color según tipo de apoyo (ya que no hay prioridad)
            if solicitud.tipo_apoyo == 'escalera':
                color = '#dc3545'  # rojo
            elif solicitud.tipo_apoyo == 'equipos':
                color = '#ffc107'  # amarillo
            else:
                color = '#6c757d'  # gris
            
            solicitudes.append({
                'id': solicitud.id,
                'tecnico_nombre': tecnico.get_nombre_completo() if tecnico else 'Desconocido',
                'coordenadas': {
                    'lat': float(solicitud.latitud),
                    'lng': float(solicitud.longitud)
                },
                'tipo_apoyo': solicitud.tipo_apoyo,
                'prioridad': 'media',  # Campo no disponible, usar valor por defecto
                'estado': solicitud.estado,
                'color': color,
                'descripcion': solicitud.observaciones or 'Sin descripción',
                'created_at': solicitud.created_at.strftime('%H:%M:%S')
            })
        
        return jsonify({
            'success': True,
            'solicitudes': solicitudes
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@lider_bp.route('/api/servicios-tiempo-real')
@login_required
@role_required('lider')
def api_servicios_tiempo_real():
    """API para obtener servicios en tiempo real para el mapa del líder"""
    try:
        servicios = []
        
        # Consultar servicios activos
        servicios_query = Servicio.query.filter(
            Servicio.estado_servicio.in_(['aceptado', 'en_ruta', 'en_sitio'])
        ).all()
        
        for servicio in servicios_query:
            # Obtener información del móvil y técnico
            movil = Usuario.query.get(servicio.movil_id) if servicio.movil_id else None
            solicitud = Solicitud.query.get(servicio.solicitud_id) if servicio.solicitud_id else None
            tecnico = Usuario.query.get(solicitud.tecnico_id) if solicitud and solicitud.tecnico_id else None
            
            # Obtener ubicación del móvil
            ubicacion_movil = None
            if movil:
                ubicacion_movil = Ubicacion.query.filter_by(
                    usuario_id=movil.id,
                    activa=True
                ).order_by(Ubicacion.timestamp.desc()).first()
            
            # Determinar color según estado
            if servicio.estado_servicio == 'aceptado':
                estado_color = '#ffc107'  # amarillo
            elif servicio.estado_servicio == 'en_ruta':
                estado_color = '#17a2b8'  # azul
            else:  # en_sitio
                estado_color = '#28a745'  # verde
            
            # Calcular tiempo de servicio
            tiempo_servicio = 0
            if servicio.aceptado_at:
                tiempo_servicio = int((datetime.utcnow() - servicio.aceptado_at).total_seconds() / 60)
            
            servicio_data = {
                'id': servicio.id,
                'movil_id': movil.id if movil else None,
                'movil_nombre': movil.get_nombre_completo() if movil else 'Desconocido',
                'tecnico_nombre': tecnico.get_nombre_completo() if tecnico else 'Desconocido',
                'estado': servicio.estado_servicio,
                'estado_color': estado_color,
                'tiempo_servicio_min': tiempo_servicio,
                'tipo_apoyo': solicitud.tipo_apoyo if solicitud else 'Desconocido'
            }
            
            # Agregar coordenadas si hay ubicación del móvil
            if ubicacion_movil:
                servicio_data['coordenadas'] = {
                    'lat': float(ubicacion_movil.latitud),
                    'lng': float(ubicacion_movil.longitud)
                }
            elif solicitud:
                # Usar coordenadas de la solicitud como fallback
                servicio_data['coordenadas'] = {
                    'lat': float(solicitud.latitud),
                    'lng': float(solicitud.longitud)
                }
            
            servicios.append(servicio_data)
        
        return jsonify({
            'success': True,
            'servicios': servicios
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@lider_bp.route('/dashboard/metricas')
@login_required
@role_required('lider')
def dashboard_metricas():
    """API para obtener métricas actualizadas del dashboard"""
    try:
        # Calcular métricas en tiempo real
        hoy = datetime.utcnow().date()
        
        # Solicitudes de hoy
        inicio_dia = datetime.combine(hoy, datetime.min.time())
        fin_dia = datetime.combine(hoy, datetime.max.time())
        solicitudes_hoy = Solicitud.query.filter(
            Solicitud.created_at.between(inicio_dia, fin_dia)
        ).count()
        
        # Servicios activos
        servicios_activos = Servicio.query.filter(
            Servicio.estado_servicio.in_(['aceptado', 'en_ruta', 'en_sitio'])
        ).count()
        
        # Tiempo promedio de respuesta (últimos 7 días)
        hace_semana = datetime.utcnow() - timedelta(days=7)
        servicios_completados = Servicio.query.filter(
            Servicio.estado_servicio == 'completado',
            Servicio.aceptado_at >= hace_semana,
            Servicio.finalizado_at.isnot(None)
        ).all()
        
        if servicios_completados:
            tiempos = []
            for servicio in servicios_completados:
                if servicio.aceptado_at and servicio.finalizado_at:
                    tiempo = (servicio.finalizado_at - servicio.aceptado_at).total_seconds() / 60
                    tiempos.append(tiempo)
            tiempo_promedio = int(sum(tiempos) / len(tiempos)) if tiempos else 0
        else:
            tiempo_promedio = 0
        
        # Tasa de aceptación (últimos 7 días)
        total_solicitudes = Solicitud.query.filter(
            Solicitud.created_at >= hace_semana
        ).count()
        
        solicitudes_aceptadas = Solicitud.query.filter(
            Solicitud.created_at >= hace_semana,
            Solicitud.estado == 'aceptada'
        ).count()
        
        tasa_aceptacion = int((solicitudes_aceptadas / total_solicitudes) * 100) if total_solicitudes > 0 else 0
        
        # Servicios activos detalle
        servicios_activos_detalle = Servicio.query.filter(
            Servicio.estado_servicio.in_(['aceptado', 'en_ruta', 'en_sitio'])
        ).all()
        
        return jsonify({
            'success': True,
            'solicitudes_hoy': solicitudes_hoy,
            'servicios_activos': servicios_activos,
            'tiempo_promedio_respuesta': tiempo_promedio,
            'tasa_aceptacion': tasa_aceptacion,
            'servicios_activos_detalle': [{
                'id': s.id,
                'estado': s.estado_servicio
            } for s in servicios_activos_detalle],
            'timestamp': datetime.utcnow().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al obtener métricas: {str(e)}'
        }), 500

@lider_bp.route('/servicios/<int:servicio_id>/reporte', methods=['POST'])
@login_required
@role_required('lider')
def generar_reporte_servicio(servicio_id):
    """Generar reporte PDF de un servicio específico"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from io import BytesIO
        
        servicio = Servicio.query.get_or_404(servicio_id)
        
        # Crear PDF en memoria
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        # Título
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, 750, f"Reporte de Servicio #{servicio.id}")
        
        # Información del servicio
        p.setFont("Helvetica", 12)
        y = 700
        
        info = [
            f"Estado: {servicio.estado_servicio}",
            f"Tipo: {servicio.solicitud.tipo_apoyo}",
            f"Dirección: {servicio.solicitud.direccion}",
            f"Coordenadas: {servicio.solicitud.latitud}, {servicio.solicitud.longitud}",
            f"Móvil: {servicio.movil.nombre if servicio.movil else 'No asignado'}",
            f"Fecha creación: {servicio.solicitud.created_at.strftime('%d/%m/%Y %H:%M')}",
            f"Fecha aceptación: {servicio.aceptado_at.strftime('%d/%m/%Y %H:%M') if servicio.aceptado_at else 'N/A'}",
            f"Fecha finalización: {servicio.finalizado_at.strftime('%d/%m/%Y %H:%M') if servicio.finalizado_at else 'N/A'}"
        ]
        
        for line in info:
            p.drawString(50, y, line)
            y -= 20
        
        # Observaciones
        if servicio.observaciones:
            y -= 20
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, "Observaciones:")
            y -= 20
            p.setFont("Helvetica", 10)
            # Dividir texto largo en líneas
            lines = [servicio.observaciones[i:i+80] for i in range(0, len(servicio.observaciones), 80)]
            for line in lines:
                p.drawString(50, y, line)
                y -= 15
        
        p.save()
        buffer.seek(0)
        
        return buffer.getvalue(), 200, {
            'Content-Type': 'application/pdf',
            'Content-Disposition': f'attachment; filename=reporte_servicio_{servicio_id}.pdf'
        }
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al generar reporte: {str(e)}'
        }), 500

@lider_bp.route('/servicios/exportar')
@login_required
@role_required('lider')
def exportar_servicios():
    """Exportar servicios a Excel"""
    try:
        import pandas as pd
        from io import BytesIO
        
        # Obtener filtros
        estado = request.args.get('estado')
        movil_id = request.args.get('movil_id')
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        
        # Construir query
        query = Servicio.query.join(Solicitud).join(Usuario, Servicio.movil_id == Usuario.id, isouter=True)
        
        if estado:
            query = query.filter(Servicio.estado_servicio == estado)
        if movil_id:
            query = query.filter(Servicio.movil_id == movil_id)
        if fecha_desde:
            query = query.filter(Servicio.aceptado_at >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
        if fecha_hasta:
            query = query.filter(Servicio.aceptado_at <= datetime.strptime(fecha_hasta, '%Y-%m-%d'))
        
        servicios = query.all()
        
        # Preparar datos para Excel
        data = []
        for servicio in servicios:
            data.append({
                'ID': servicio.id,
                'Estado': servicio.estado_servicio,
                'Tipo': servicio.solicitud.tipo_apoyo,
                'Coordenadas': f"{servicio.solicitud.latitud}, {servicio.solicitud.longitud}",
                'Dirección': servicio.solicitud.direccion,
                'Móvil': servicio.movil.nombre if servicio.movil else 'No asignado',
                'Fecha Creación': servicio.solicitud.created_at.strftime('%d/%m/%Y %H:%M'),
                'Fecha Aceptación': servicio.aceptado_at.strftime('%d/%m/%Y %H:%M') if servicio.aceptado_at else '',
                'Fecha Finalización': servicio.finalizado_at.strftime('%d/%m/%Y %H:%M') if servicio.finalizado_at else '',
                'Observaciones': servicio.observaciones or ''
            })
        
        # Crear DataFrame y Excel
        df = pd.DataFrame(data)
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Servicios', index=False)
        
        buffer.seek(0)
        
        return buffer.getvalue(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename=servicios_export_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
        }
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al exportar servicios: {str(e)}'
        }), 500

@lider_bp.route('/api/solicitudes/export')
@login_required
@role_required('lider')
def exportar_solicitudes():
    """Exportar solicitudes a Excel con filtros opcionales"""
    try:
        # Obtener parámetros de filtro
        estado = request.args.get('estado')
        tecnico_id = request.args.get('tecnico_id')
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        tipo_apoyo = request.args.get('tipo_apoyo')
        localidad_codigo = request.args.get('localidad')
        
        # Construir query base
        query = db.session.query(Solicitud).join(Usuario, Solicitud.tecnico_id == Usuario.id)
        
        # Aplicar filtros
        if estado:
            query = query.filter(Solicitud.estado == estado)
        if tecnico_id:
            query = query.filter(Solicitud.tecnico_id == tecnico_id)
        if tipo_apoyo:
            query = query.filter(Solicitud.tipo_apoyo == tipo_apoyo)
        if fecha_desde:
            query = query.filter(Solicitud.created_at >= datetime.strptime(fecha_desde, '%Y-%m-%d'))
        if fecha_hasta:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Solicitud.created_at < fecha_hasta_dt)
        
        # Ordenar por fecha de creación descendente
        query = query.order_by(Solicitud.created_at.desc())
        
        # Obtener solicitudes base
        solicitudes_base = query.all()
        
        # Filtrar por localidad si se especifica
        if localidad_codigo and localidad_codigo != 'todas':
            localidad = Localidad.get_by_codigo(localidad_codigo)
            if localidad:
                solicitudes = []
                for solicitud in solicitudes_base:
                    if solicitud.latitud and solicitud.longitud:
                        if localidad.contains_point(float(solicitud.latitud), float(solicitud.longitud)):
                            solicitudes.append(solicitud)
            else:
                solicitudes = []
        else:
            solicitudes = solicitudes_base
        
        # Preparar datos para Excel
        data = []
        for solicitud in solicitudes:
            # Obtener información del servicio asociado si existe
            servicio = solicitud.servicio
            
            data.append({
                'ID Solicitud': solicitud.id,
                'Estado': solicitud.estado.title(),
                'Tipo de Apoyo': solicitud.tipo_apoyo.title(),
                'Técnico': solicitud.tecnico.get_nombre_completo(),
                'Teléfono Técnico': solicitud.tecnico.telefono or '',
                'Coordenadas': f"{solicitud.latitud}, {solicitud.longitud}",
                'Dirección': solicitud.direccion or '',
                'Observaciones': solicitud.observaciones or '',
                'Fecha Creación': solicitud.created_at.strftime('%d/%m/%Y %H:%M'),
                'Límite Tiempo': solicitud.limite_tiempo.strftime('%d/%m/%Y %H:%M'),
                'Última Actualización': solicitud.updated_at.strftime('%d/%m/%Y %H:%M') if solicitud.updated_at else '',
                'ID Servicio': servicio.id if servicio else '',
                'Estado Servicio': servicio.estado_servicio.title() if servicio else '',
                'Móvil Asignado': servicio.movil.nombre if servicio and servicio.movil else '',
                'Fecha Aceptación': servicio.aceptado_at.strftime('%d/%m/%Y %H:%M') if servicio and servicio.aceptado_at else '',
                'Fecha Finalización': servicio.finalizado_at.strftime('%d/%m/%Y %H:%M') if servicio and servicio.finalizado_at else '',
                'Duración (min)': servicio.duracion_minutos if servicio else '',
                'Observaciones Finales': servicio.observaciones_finales if servicio else ''
            })
        
        # Crear DataFrame y Excel
        df = pd.DataFrame(data)
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Solicitudes', index=False)
            
            # Obtener el workbook y worksheet para formatear
            workbook = writer.book
            worksheet = writer.sheets['Solicitudes']
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'solicitudes_export_{timestamp}.xlsx'
        
        return buffer.getvalue(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename={filename}'
        }
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al exportar solicitudes: {str(e)}'
        }), 500

# ==================== RUTAS DE GESTIÓN DE VEHÍCULOS ====================

@lider_bp.route('/vehiculos')
@login_required
@role_required('lider')
def vehiculos():
    """Lista todos los vehículos"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = 10
        
        # Filtros
        search = request.args.get('search', '').strip()
        tipo_vehiculo = request.args.get('tipo_vehiculo', '')
        estado_asignacion = request.args.get('estado_asignacion', '')
        
        # Query base
        query = Vehiculo.query.filter_by(activo=True)
        
        # Aplicar filtros
        if search:
            query = query.filter(
                or_(
                    Vehiculo.placa.ilike(f'%{search}%'),
                    Vehiculo.marca.ilike(f'%{search}%'),
                    Vehiculo.modelo.ilike(f'%{search}%')
                )
            )
        
        if tipo_vehiculo:
            query = query.filter_by(tipo_vehiculo=tipo_vehiculo)
        
        if estado_asignacion == 'asignado':
            query = query.filter(Vehiculo.movil_id.isnot(None))
        elif estado_asignacion == 'sin_asignar':
            query = query.filter(Vehiculo.movil_id.is_(None))
        
        # Ordenar por fecha de creación descendente
        query = query.order_by(Vehiculo.created_at.desc())
        
        # Paginación
        vehiculos_paginados = query.paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # Obtener móviles disponibles para asignación
        moviles_disponibles = Usuario.query.filter_by(
            rol='movil', 
            activo=True
        ).order_by(Usuario.nombre, Usuario.apellido).all()
        
        return render_template('lider/vehiculos.html',
                             vehiculos=vehiculos_paginados,
                             moviles_disponibles=moviles_disponibles,
                             search=search,
                             tipo_vehiculo=tipo_vehiculo,
                             estado_asignacion=estado_asignacion)
                             
    except Exception as e:
        flash(f'Error al cargar vehículos: {str(e)}', 'error')
        return redirect(url_for('lider.dashboard'))

@lider_bp.route('/vehiculos/nuevo', methods=['GET', 'POST'])
@login_required
@role_required('lider')
def nuevo_vehiculo():
    """Crear un nuevo vehículo"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            placa = request.form.get('placa', '').strip().upper()
            marca = request.form.get('marca', '').strip()
            modelo = request.form.get('modelo', '').strip()
            año = request.form.get('año', type=int)
            color = request.form.get('color', '').strip()
            tipo_vehiculo = request.form.get('tipo_vehiculo', '').strip()
            combustible = request.form.get('combustible', '').strip()
            numero_motor = request.form.get('numero_motor', '').strip() or None
            numero_chasis = request.form.get('numero_chasis', '').strip() or None
            cilindraje = request.form.get('cilindraje', '').strip() or None
            observaciones = request.form.get('observaciones', '').strip() or None
            movil_id = request.form.get('movil_id', type=int) or None
            
            # Validaciones
            if not all([placa, marca, modelo, año, color, tipo_vehiculo, combustible]):
                flash('Todos los campos obligatorios deben ser completados.', 'error')
                return redirect(url_for('lider.nuevo_vehiculo'))
            
            # Verificar que la placa no exista
            vehiculo_existente = Vehiculo.query.filter_by(placa=placa).first()
            if vehiculo_existente:
                flash(f'Ya existe un vehículo con la placa {placa}.', 'error')
                return redirect(url_for('lider.nuevo_vehiculo'))
            
            # Verificar que el móvil no tenga otro vehículo asignado
            if movil_id:
                vehiculo_movil = Vehiculo.query.filter_by(movil_id=movil_id, activo=True).first()
                if vehiculo_movil:
                    movil = Usuario.query.get(movil_id)
                    flash(f'El móvil {movil.get_nombre_completo()} ya tiene un vehículo asignado.', 'error')
                    return redirect(url_for('lider.nuevo_vehiculo'))
            
            # Crear el vehículo
            vehiculo = Vehiculo(
                placa=placa,
                marca=marca,
                modelo=modelo,
                año=año,
                color=color,
                tipo_vehiculo=tipo_vehiculo,
                combustible=combustible,
                numero_motor=numero_motor,
                numero_chasis=numero_chasis,
                cilindraje=cilindraje,
                observaciones=observaciones,
                movil_id=movil_id
            )
            
            db.session.add(vehiculo)
            db.session.commit()
            
            flash(f'Vehículo {placa} creado exitosamente.', 'success')
            return redirect(url_for('lider.vehiculos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear vehículo: {str(e)}', 'error')
            return redirect(url_for('lider.nuevo_vehiculo'))
    
    # GET - Mostrar formulario
    moviles_disponibles = Usuario.query.filter_by(
        rol='movil', 
        activo=True
    ).order_by(Usuario.nombre, Usuario.apellido).all()
    
    return render_template('lider/vehiculo_form.html',
                         vehiculo=None,
                         moviles_disponibles=moviles_disponibles,
                         titulo='Nuevo Vehículo')

@lider_bp.route('/vehiculos/<int:vehiculo_id>/editar', methods=['GET', 'POST'])
@login_required
@role_required('lider')
def editar_vehiculo(vehiculo_id):
    """Editar un vehículo existente"""
    vehiculo = Vehiculo.query.get_or_404(vehiculo_id)
    
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            placa = request.form.get('placa', '').strip().upper()
            marca = request.form.get('marca', '').strip()
            modelo = request.form.get('modelo', '').strip()
            año = request.form.get('año', type=int)
            color = request.form.get('color', '').strip()
            tipo_vehiculo = request.form.get('tipo_vehiculo', '').strip()
            combustible = request.form.get('combustible', '').strip()
            numero_motor = request.form.get('numero_motor', '').strip() or None
            numero_chasis = request.form.get('numero_chasis', '').strip() or None
            cilindraje = request.form.get('cilindraje', '').strip() or None
            observaciones = request.form.get('observaciones', '').strip() or None
            movil_id = request.form.get('movil_id', type=int) or None
            
            # Validaciones
            if not all([placa, marca, modelo, año, color, tipo_vehiculo, combustible]):
                flash('Todos los campos obligatorios deben ser completados.', 'error')
                return redirect(url_for('lider.editar_vehiculo', vehiculo_id=vehiculo_id))
            
            # Verificar que la placa no exista en otro vehículo
            vehiculo_existente = Vehiculo.query.filter(
                Vehiculo.placa == placa,
                Vehiculo.id != vehiculo_id
            ).first()
            if vehiculo_existente:
                flash(f'Ya existe otro vehículo con la placa {placa}.', 'error')
                return redirect(url_for('lider.editar_vehiculo', vehiculo_id=vehiculo_id))
            
            # Verificar que el móvil no tenga otro vehículo asignado
            if movil_id and movil_id != vehiculo.movil_id:
                vehiculo_movil = Vehiculo.query.filter_by(movil_id=movil_id, activo=True).first()
                if vehiculo_movil:
                    movil = Usuario.query.get(movil_id)
                    flash(f'El móvil {movil.get_nombre_completo()} ya tiene un vehículo asignado.', 'error')
                    return redirect(url_for('lider.editar_vehiculo', vehiculo_id=vehiculo_id))
            
            # Actualizar el vehículo
            vehiculo.placa = placa
            vehiculo.marca = marca
            vehiculo.modelo = modelo
            vehiculo.año = año
            vehiculo.color = color
            vehiculo.tipo_vehiculo = tipo_vehiculo
            vehiculo.combustible = combustible
            vehiculo.numero_motor = numero_motor
            vehiculo.numero_chasis = numero_chasis
            vehiculo.cilindraje = cilindraje
            vehiculo.observaciones = observaciones
            vehiculo.movil_id = movil_id
            
            db.session.commit()
            
            flash(f'Vehículo {placa} actualizado exitosamente.', 'success')
            return redirect(url_for('lider.vehiculos'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar vehículo: {str(e)}', 'error')
            return redirect(url_for('lider.editar_vehiculo', vehiculo_id=vehiculo_id))
    
    # GET - Mostrar formulario
    moviles_disponibles = Usuario.query.filter_by(
        rol='movil', 
        activo=True
    ).order_by(Usuario.nombre, Usuario.apellido).all()
    
    return render_template('lider/vehiculo_form.html',
                         vehiculo=vehiculo,
                         moviles_disponibles=moviles_disponibles,
                         titulo='Editar Vehículo')

@lider_bp.route('/vehiculos/<int:vehiculo_id>/eliminar', methods=['POST'])
@login_required
@role_required('lider')
def eliminar_vehiculo(vehiculo_id):
    """Eliminar (desactivar) un vehículo"""
    try:
        vehiculo = Vehiculo.query.get_or_404(vehiculo_id)
        
        # Desactivar en lugar de eliminar
        vehiculo.activo = False
        vehiculo.movil_id = None  # Desasignar móvil
        
        db.session.commit()
        
        flash(f'Vehículo {vehiculo.placa} eliminado exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar vehículo: {str(e)}', 'error')
    
    return redirect(url_for('lider.vehiculos'))

@lider_bp.route('/vehiculos/<int:vehiculo_id>/asignar', methods=['POST'])
@login_required
@role_required('lider')
def asignar_vehiculo(vehiculo_id):
    """Asignar un móvil a un vehículo"""
    try:
        vehiculo = Vehiculo.query.get_or_404(vehiculo_id)
        movil_id = request.form.get('movil_id', type=int)
        
        if not movil_id:
            flash('Debe seleccionar un móvil.', 'error')
            return redirect(url_for('lider.vehiculos'))
        
        # Verificar que el móvil existe y es activo
        movil = Usuario.query.filter_by(id=movil_id, rol='movil', activo=True).first()
        if not movil:
            flash('El móvil seleccionado no es válido.', 'error')
            return redirect(url_for('lider.vehiculos'))
        
        # Verificar que el móvil no tenga otro vehículo asignado
        vehiculo_existente = Vehiculo.query.filter_by(movil_id=movil_id, activo=True).first()
        if vehiculo_existente and vehiculo_existente.id != vehiculo_id:
            flash(f'El móvil {movil.get_nombre_completo()} ya tiene un vehículo asignado.', 'error')
            return redirect(url_for('lider.vehiculos'))
        
        # Asignar el vehículo
        vehiculo.movil_id = movil_id
        db.session.commit()
        
        flash(f'Vehículo {vehiculo.placa} asignado a {movil.get_nombre_completo()} exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al asignar vehículo: {str(e)}', 'error')
    
    return redirect(url_for('lider.vehiculos'))

@lider_bp.route('/vehiculos/<int:vehiculo_id>/desasignar', methods=['POST'])
@login_required
@role_required('lider')
def desasignar_vehiculo(vehiculo_id):
    """Desasignar un móvil de un vehículo"""
    try:
        vehiculo = Vehiculo.query.get_or_404(vehiculo_id)
        
        if not vehiculo.movil_id:
            flash('El vehículo no tiene un móvil asignado.', 'error')
            return redirect(url_for('lider.vehiculos'))
        
        movil_nombre = vehiculo.get_movil_asignado()
        vehiculo.movil_id = None
        
        db.session.commit()
        
        flash(f'Vehículo {vehiculo.placa} desasignado de {movil_nombre} exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al desasignar vehículo: {str(e)}', 'error')
    
    return redirect(url_for('lider.vehiculos'))

@lider_bp.route('/api/vehiculos')
@login_required
@role_required('lider')
def api_vehiculos():
    """API para obtener lista de vehículos"""
    try:
        vehiculos = Vehiculo.query.filter_by(activo=True).order_by(Vehiculo.placa).all()
        
        return jsonify({
            'success': True,
            'vehiculos': [vehiculo.to_dict() for vehiculo in vehiculos]
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al obtener vehículos: {str(e)}'
        }), 500